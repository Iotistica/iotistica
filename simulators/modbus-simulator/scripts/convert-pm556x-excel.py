#!/usr/bin/env python3
"""
Convert Schneider PM556x Excel register list to Modbus simulator JSON profile.

Usage:
    python convert-pm556x-excel.py

Requires:
    pip install openpyxl
"""

import openpyxl
import json
import os

# Path to Excel file
EXCEL_FILE = "profiles/PM556x_PublicModbusRegisterList_10th_August-23.xlsx"
OUTPUT_FILE = "profiles/PM556x.json"

# Filter to specific categories (empty list = all categories)
INCLUDE_CATEGORIES = ["Meter Data (Basic)"]  # Only include these categories

def parse_excel_to_profile():
    """Parse Excel file and convert to Modbus simulator profile format"""
    
    print(f"Loading Excel file: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # Show all available sheets
    print(f"\nAvailable sheets: {wb.sheetnames}")
    
    # Use the "Register List" sheet specifically
    sheet = wb["Register List"]
    print(f"Using sheet: {sheet.title}")
    
    # Display first 10 rows to understand structure
    print("\nFirst 10 rows (for understanding structure):")
    for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
        print(f"Row {i}: {row}")
    
    print("\n" + "="*80)
    print("Please examine the rows above and update this script with:")
    print("1. Which row contains column headers?")
    print("2. Which columns contain: Address, Name, Type, Unit, etc.")
    print("="*80)
    
    # Based on Schneider PM556x Excel structure
    HEADER_ROW = 1     # Row 1 has column names
    CATEGORY_COL = 0   # Column 0 = 'Category'
    SUBCAT1_COL = 1    # Column 1 = 'Sub Cat 1'
    SUBCAT2_COL = 2    # Column 2 = 'Sub Cat 2'
    ADDRESS_COL = 15   # Column 15 = 'Register' (Modbus address)
    NAME_COL = 3       # Column 3 = 'Description' (register name)
    UNIT_COL = 16      # Column 16 = 'Units'
    DATA_TYPE_COL = 18 # Column 18 = 'Data Type' (INT16, INT32, FLOAT32, etc.)
    ACCESS_COL = 19    # Column 19 = 'Access' (R, RW, RWC, etc.)
    
    # Track current category hierarchy
    current_category = None
    current_subcat1 = None
    current_subcat2 = None
    
    # Organize by categories
    categories = {}
    
    # Skip header rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True), HEADER_ROW + 1):
        if not row or not row[ADDRESS_COL]:
            continue
        
        # Update category tracking
        if row[CATEGORY_COL]:
            current_category = str(row[CATEGORY_COL])
        if row[SUBCAT1_COL]:
            current_subcat1 = str(row[SUBCAT1_COL])
        if row[SUBCAT2_COL]:
            current_subcat2 = str(row[SUBCAT2_COL])
        
        # Skip category headers (have text in first column instead of number in address column)
        try:
            address = int(row[ADDRESS_COL])
        except (ValueError, TypeError):
            continue  # Skip non-numeric addresses (category headers)
        
        name = str(row[NAME_COL]) if row[NAME_COL] else f"Register{address}"
        unit = str(row[UNIT_COL]) if row[UNIT_COL] else ""
        access = str(row[ACCESS_COL]).upper() if row[ACCESS_COL] else "R"
        data_type = str(row[DATA_TYPE_COL]) if row[DATA_TYPE_COL] else "INT16"
        
        # Determine register type from Access column
        # R = Read-only = Input Register
        # RW, RWC = Read-Write = Holding Register
        if access.startswith('R') and 'W' not in access:
            reg_type = "input"  # Read-only
        else:
            reg_type = "holding"  # Read-write or writable
        
        try:
            # Determine base value from unit type
            base_values = {
                "V": 230,
                "A": 10,
                "W": 1000,
                "kW": 5,
                "VA": 1000,
                "kVA": 5,
                "Hz": 50,
                "PF": 95,
                "°C": 25,
                "%": 50,
                "Wh": 1000,
                "kWh": 5,
                "VAh": 1000,
                "kVAh": 5
            }
            
            base = 100  # default
            for unit_key, unit_base in base_values.items():
                if unit_key in unit:
                    base = unit_base
                    break
            
            # Adjust noise based on data type
            noise_pct = 0.02  # 2% default
            if "FLOAT" in data_type.upper():
                noise_pct = 0.01  # 1% for floating point (more precision)
            
            # Build category path
            cat_path = current_category or "Uncategorized"
            if current_subcat1:
                cat_path = f"{cat_path} > {current_subcat1}"
            if current_subcat2:
                cat_path = f"{cat_path} > {current_subcat2}"
            
            data_point = {
                "address": address,
                "type": reg_type,
                "name": name,
                "unit": unit,
                "data_type": data_type,
                "access": access,
                "base": base,
                "noise_pct": noise_pct,
                "category": cat_path
            }
            
            # Add to category-organized structure
            if current_category not in categories:
                categories[current_category] = []
            categories[current_category].append(data_point)
            
        except (IndexError, ValueError) as e:
            print(f"Skipping row {row_idx}: {e}")
            continue
    
    # Flatten to single dataPoints array (simulator expects this format)
    # Apply category filter if specified
    all_data_points = []
    for cat_name, cat_points in categories.items():
        if not INCLUDE_CATEGORIES or cat_name in INCLUDE_CATEGORIES:
            all_data_points.extend(cat_points)
    
    # Create profile
    profile = {
        "PM556x": {
            "vendor": "Schneider Electric",
            "model": "PM556x Power Meter",
            "version": "1.0.0",
            "vendorUrl": "https://www.se.com",
            "dataPoints": all_data_points
        }
    }
    
    # Save to JSON
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(profile, f, indent=2)
    
    total_points = len(all_data_points)
    print(f"\n✓ Created profile with {total_points} data points")
    print(f"✓ Saved to: {OUTPUT_FILE}")
    
    # Show breakdown by category
    print(f"\nBreakdown by category:")
    for cat_name, cat_points in sorted(categories.items(), key=lambda x: len(x[1]), reverse=True):
        print(f"  {cat_name}: {len(cat_points)} registers")
    
    # Show sample from first category
    first_cat = next(iter(categories.values()))
    print(f"\nSample registers (first 5):")
    for dp in first_cat[:5]:
        cat_short = dp['category'].split(' > ')[-1][:20]
        print(f"  {dp['address']:5d} | {dp['type']:8s} | {dp['name'][:30]:30s} | {dp['unit']:6s} | {cat_short}")

if __name__ == "__main__":
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file not found: {EXCEL_FILE}")
        print("\nPlease ensure the file exists at:")
        print(f"  {os.path.abspath(EXCEL_FILE)}")
        exit(1)
    
    parse_excel_to_profile()
